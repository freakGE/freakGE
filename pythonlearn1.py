#  https://www.youtube.com/watch?v=_uQrJ0TkZlc&list=WL&index=15&t=1417s 

# 1 Variables
'''

price = 10
rating = 4.9
name = 'Saba'
is_published = True
print(price)

'''
# 2 Receiving Input
'''

full_name = 'Saba Esebua'
age = 20
is_new = True
'''
'''
----------
inputshi vwert texts magalitad shekitxvis saxit ra gqvia shekitxvis
shemdeg ert ujreds vtovebt rata gashvebis shemdeg gavcet pasuxi
meore linshi vwert axal cvlads da inputshi axali shekitxvaa sheni
favoriti feri.
printit ki vidzaxebt name shi rasac vhawer shemdgom ubralod texts
da amis mere ki idzaxeb fers
----------
name = input('What is your name? ')# shekitxvis shemdeg ert ujreds vtovebt
favorite_color = input('What is your favorite color? ')
print(name + ' likes ' + favorite_color)

'''

# 3 Python Cheat Sheet  /    Type Conversion 
'''
---------
gvaqvs cvladi birth_year da inputshi vtxovt sheiyvanos
dabadebis tarigi, shemdeg vbechdavt type ti ki mis class gavigebt
gvaqvs axali cvladi age = wlevandel ricxvs gamoklebuli iq chawerili
dabadebis tarigi anu 2020 - 2003 chems shemtxvevashi
amjamad integer shi vwert mas, shemdeg vprintavt type is mititebit..
---------
birth_year = input('Birth year: ')
print(type(birth_year))
age = 2020 - int(birth_year)
print(type(age))
print(age)

'''
# 4

'''
--------
vutitebt lbs da inputshi vwert rom mivutitot chveni lbs..
meore cvladshi ki mivigebt pasuxs int formashi aris chawerili weight_lbs
pirveli cvladi anu iq rasac chavwert * 0.45
amit chven wonas gavigebt
--------

weight_lbs = input('Weight (lbs): ')
weight_kg = int(weight_lbs) * 0.45
print(weight_kg)

'''

# 5 Strings

'''       
print(course[0:9]) [] am frcxhilshi chawerili cifrit vprintavt mxolod im
rigshi myofs shegvidzlia kide magalitad 0 anu P dan meore sityvis bolo
asomde magalitad [0:10] am dros print gamogvitans "Python for"
xolo frchxilebshi minus nishnavs shetrialebuli formit wamova magalitad
[-2] nishnavs "r" s. aseve shegvidzlia gamovitanot [1:] am magalitit
"ython for Begginers" anu 1 dan bolomde amitomac amoagdo 0 anu P
sruli text albat xvdebit iqneboda [0:] aseve shegvidzlia piriqit
[:6] nishnavs rom amobechdos pirveli 6 aso anu "Python"
xolo damatebit kide shegvidzlia cvlads gavaigivot axal cvladtan anu
course = Python for begginers
magram another = course
anu another = python for begginers ... too
--------------

course = 'Python for Beginners'
#         0123456789   -4-3-2-1
another = course[:]
print(another)

'''
'''
am dros daprintavs pirvelis da bolo asos gareshe anu mivigebt "nastasi"
name = 'Anastasia'
print(name[1:-1])

'''

# 6 Formatted Strings
''' gaqvs cvladi first anu firstname Saba, meore last Esebua
message shi vutitebt rom first anu Saxels + Gvari da text is a coder
msg cvladshi ki ufro kargad aris dawerili magram programis
gashvebisas analogiuria orive message da msg
magram jobs msg stilshi davwerot shemdegshi
-------
first = 'Saba'
last = 'Esebua'
message = first + ' [' + last + '] is a coder'
msg = f'{first} [{last}] is a coder'

print(msg)
print(message)

'''

# 7 Strings Methods
'''
course = 'Python for Beginners'
print(len(course)) # len  anu course ramdeni aso aris..
'''
'''
course = 'Python for Beginners'
print(course.upper()) # .upper anu yvela aso didi asoti wamoachinos textshi
'''
'''
course = 'Python for Beginners'
print(course.lower()) # lower yvela aso patara unda warmoachinos.
'''
'''
----------
es ki yvela zevit mocemuli magaliti ertad isini yvela calcalke moqmedebs
anu tu iq course.lower an upperis emre vizavt print course
course arsheicvleba da course = 'Python for Beginners' iqneba
tavdapirveli cvladis texti
----------
course = 'Python for Beginners'
print(len(course))
print(course.upper())
print(course.lower())
print(course)
'''
'''
course = 'Python for Beginners'
print(course.find('P')) # pasuxi iqneba 0 radgan fin modzebnis meramdenea.
print(course.find('0')) # pasuxi iqneba -1
print(course.find('o')) # pasuxi iqneba 4 randgan pirvelad 0 gxvdeba meotxe asod
print(course.find('Beginners')) # = 11 B sityvis pirveli aso daiwyo 11dan.
print(course.replace('Beginners', 'Absolute Begginers'))
'Python for Beginners' in course
# cvehn replace ti Beginners chavanacvlet Absolute Begginerit
# da texts mivigebt Python for Absolute Beginners
print(course.replace('P', 'J')) #
len() 

'''
# 8 Arithmetic Operations 
'''
x = 10
x -= 3
x += 3
print(x)
print(10 + 3)
print(9 / 3)
print(10 // 3)
print(10 * 10)

'''

# 9 Operator Precedence
'''
x = (10 + 3) * 2 ** 2
print(x)
# parenthesis
# exponentiation 2 ** 3
# multiplication or division
# addition or subtraction
'''

# 10 Math Functions
'''
import math

#print(math.floor(2.9)) # anu wina cifrs amobechdavs anu am shetxvevashi 2's
#print(math.ceil(2.9)) 
x = 2.9
print(round(x)) # damrgvaleba
#print(abs(-2.9))
'''

# 11 If Statments
'''

is_cold = False
is_hot = False

if is_cold:
    print("Dges dzalian civa")
    print("cxeli shokoladi gaasworebda")
elif is_hot:    
    print("Dges dzalian cxela")
    print("Civi yava gaasworebda")
else:
    print("Dges kargi dgea")

print("- Ho martali xar")

'''
'''
price = 1000000
has_good_credit = True

if has_good_credit:
    down_payment = 0.1 * price
else:
    down_payment = 0.2 * price
print(f"Down payment: ${down_payment}")
'''

# 12 Logical Operators

# AND = orive
# or = Minimum erti
# not 
'''
saxlshi_mivdivar = True
kargi_krediti = False

if saxlshi_mivdivar or kargi_krediti: # [ or / AND ] 
    print("Saxshi asworebs")
'''
'''
kargi_ambavia = True
kriminali_daichires = False

if kargi_ambavia and not kriminali_daichires: # anu tu meore cvlads false aqvs
    print("Saxlshi asworebs")         # programa imushavbs tuarada ar imushavebs

'''
# 13 Comparison Operatos
'''
temperature = 35

if temperature != 30: # <= / >=  /  ==  / > / < / != /
    print("Dges cxela")
else:
    print("Dges sicxe araa")

'''
'''
#name = "Saba Esebua" # cvladi
name = input("Saxeli: ")

if len(name) < 3: #len gamoitvlis tu ramdeni characteria textshi anu tu 3ze 
    print("Saxelshi minimum 3 aso unda iyos") # naklebia daprintos es
elif len(name) > 50: # tu 50 charracterze metia daprintos es
    print("Saxelshi maqsimum 50 aso sheileba iyos")
else: # tu 3 ze metia da 50 ze naklebi daprintos kargi saxelia..
    print("Kargi saxelia!")

'''
# 14 Project: Weight Converter
'''
weight = int(input('Weight: ')) # cvladi int formashi
unit = input('(L)bs or (K)g: ') # meore cvladi..
if unit.upper() == "L": # tu "L" s davachert lbs dan kiloze gadava
    converted = weight * 0.45
    print(f"You are {converted} kilos")
else: # tu k s avirchevt unit shi anu gavigebt ramdeni lbs vart
    converted = weight / 0.45
    print(f"You are {converted} pounds")
'''
# 15 While Loops
'''
i = 1 # cvladi i = 1 
while i <= 5: # xolo "i" tu  naklebia an toli 5ze daprintos * gamravlebuli ize
   # print(i) # tu mxolod cifri gvinda print(i)
    print('*' * i)# i = i + 1 anu nelnela daumatos mas fifqebi axal lainze da 
    i = i + 1  # dasurlebis shemdeg daweros "Done"
print("Done")    
'''

# 16 Building a Guessing Game  / Gamocnoba..
'''
secret_number = 9 # cvladi 9 tu gamoicnob am ricvxs gaimarjev anu..
guess_count = 0 
guess_limit = 3 # limit am tamashis gamocnobis aris 3 shansi
while guess_count < guess_limit: # tu gues count naklebia 3 ze daprintos es
    guess = int(input('Gamoicani: ')) 
    guess_count += 1 # yovel cdaze rom agiricxos += 1 anu 3 jer tu qna waagebs
    if guess == secret_number: # tu kitxvashi chawers 9s daprintavs rom
        print('Tqven gaimarjvet!') # man moigo
        break # tu gamocnobs programa shewydeba
else: # tu sami shansis mere waago printavs amas
    print('Samwuxarod tqven waaget')
'''
# 17 Building the Car Game

'''
command = ""
while command.lower() != "quit":
    command = input("> ")
    if.command.lower() == "start":
        print("Car started...")
    elif command.lower() == "stop":
        print("Car stopped.")
# tu gvinda rom command.lower() ar gavimeorot bevrjer
# mogviwevs qvevit mocemuli magalitis mesame xazshi chawerili
# kodis dawera anu boloshi mivutito ".lower()
# vutitebt im commands anu amieridan sul ase gaishvebs programa
# patara anu lower shriftit
'''
'''
command = ""
started = False
while True:
# while command != "quit": aseve esec sheidzleba while trues nacvlad
    command = input("> ").lower()
    if command == "start":
        if started:
            print("Car is already started")
        else:
            started = True
            print("Car started...")
    elif command == "stop":
        if not started:
            print("Car is already stopped!")
        else:
            started = False
            print("Car stopped.")
    elif command == "help":
        print("""
start - to start the car
stop - to stop the car
quit - to quit
        """)
    elif command =="quit":
        break
    else:
        print("Sorry, I don't understand that!")
'''
# 18 For Loops
'''
for item in ['Saba', 'Vaso', 'Nia']:
    print(item)
# tu gvinda rom cifrebi daweros vwer rangeshi magalitad (100)
for items in range(10, 105, 5):  # anu atidan 105mde xutxutis dashorebit
    print(items)
'''
'''
prices = [10, 20, 30]

total = 0
for price in prices:
    total += price
print(f"Total: {total}")    
'''
# 19 Nested Loops
'''
for x in range(4):
    for y in range(3):
        print(f'({x}, {y})')
'''
'''
numbers = [5, 2, 5, 2, 2]
for x_count in numbers:   
#    print('x' * x_count)  # shegvidzlia martivad amit davwerot mxolod
    output = ''
    for count in range(x_count):
        output += 'x'
    print(output)
'''
# 20 Lists
'''
names = ['Saba', 'Gio', 'Mari', 'Tekla', 'Sofo']
#          0       1      2        3        4
#         -5      -4     -3       -2       -1
names[0] = 'Sabunia' # amit sabas chavancvlebt sabuniashi update..
print(names[2:4])
print(names[1:])
print(names[-3])
print(names)
'''
''' 
# am kodit chven numbershi yvelaze magal ricxvs gviprintavs
numbers = [3, 6, 2, 8, 4, 10]
max = numbers[0]
for number in numbers:
    if number > max:
        max = number
print(max)
'''

# 21    2D Lists
'''
matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
#matrix[0][1] = 20  axla ki matrix shegvidzlia im cifris 
# nacvlad 20 chavwerot
#print(matrix[0][1]) amit daprintavs 2's
for row in matrix:
    for item in row:
        print(item)
# chven yvela ricxvi chamovweret tu raciyo matrixshi
'''
# 22 List Methods
'''
numbers = [5, 2, 1, 7, 4]
#numbers.append(20) amatebs cvlads 
#numbers.remove(5) remove ti shlis im characters
#numbers.clear() yvelafers shlis im cvladidan yvela characters
#numbers.pop() shlis bolos damatebul cvlads anu amjamad 4's
#print(numbers.index(5))
#print(numbers.count(5))#gavigebt ramdenjeraa chawerili anu aq 1xel
#numbers.sort() # es mciredan didi cifrisken midis da mwrkivshi ayenebs
#numbers.reverse() # reverse abrunebs maglidan qvevit..
numbers2 = numbers.copy() # akopirebs cvlads.. da shignit yvelafers
numbers.append(10) # magram rodesac pirvel cvladshi davamatebt
# numbers 2 printvisdros iqneba im droindeli number pirveli cvladi
# da arnishnavs ro amieridan ertnairi mnishvnelobis cvladi iqneba
print(numbers2)
'''
'''
numbers = [2, 2, 4, 6, 3, 4, 6, 1]
uniques = []
for number in numbers: # anu orjer chaweril ricxvebi ar daiprintos
    if number not in uniques: # amistvis vaketebt msgavs scripts
        uniques.append(number) 
print(uniques)
'''
# 23 Tuples
''' tu programas gavushvebt tupleti anu msgavsi prchyalebit..
chveulebrivad shegvidzlia davprintot misi romelime wevri da sxva yvelaferi
magram arshegvdzizlia magalitad shevcvalot '''
'''
numbers = (1, 2, 3)
# numbers[0] = 10 arshegvdizlia 1 shevcvalot 10d anu..
print(numbers[0])
'''
#24 Unpacking
''' 
coordinates = (1, 2, 3)
# coordinates = [1, 2, 3] tu programas gavushvebt msgavsi frcxhilebit
# mivigebt mainc rom x udris coordinates 1's 
#x = coordinates[0]
#y = coordinates[1]
#z = coordinates[2]
x, y, z = coordinates
print(x)
'''
#25 Dictionaries
'''
customer = {
    "name": "Saba Esebua",
    "age": 16,
    "is_verifed": True
}
#print(customer["name"]) # tu magalitad Name chavwer mivigebt errors
#print(customer.get("birthdate", "Jum 9 2003")) 
'''
'''
 tu birtdays chavwert mxolod
programa naxavs rom zevit araris chawerili birtday da dagviwers None's
magram mdzimes shemdeg chven vutitebt tu ragvinda iq ro chavwerot
amistvis gvexmareba customer.get "get" '''
'''
customer["name"] = "Nini Esebua" # am metodit martivad vamatebt msgavs moqmedebebs
print(customer["name"])
'''
'''
phone = input("Phone: ")
digits_mapping = {
    "1": "One",
    "2": "Two",
    "3": "Three",
    "4": "Four"
}
output = ""
for ch in phone:
    output += digits_mapping.get(ch, "!") + " "
print(output)    
'''
# 26 Emoji Converter :)
'''
message = input(">")
words = message.split(' ')
#Good morning :)

emojis = {
    ":)": "\N{winking face}",
}
output = ""
for word in words:
    output += emojis.get(word, word) + " "
print(output)   
'''
# 27 Functions
'''
def it vqmnit funqcias romelic moxseniebis dros daprintavs mas 
rogorc bolo sam lains xedavt greet user aris shuashi
amitom textis dros pirveli iqneba start shemdeg greetusershi chawerili
shemdeg ki finish
#greet_user() tu zevit movixseniebt greetuser erros amoagdebs
def greet_user(): 
    print('Hi there!')
    print('Welcome aboard') 


print("Start")
greet_user()
print("Finish")
'''
#28 Parameters
'''
def greet_user(name):  
    print(f'Hi {name}!')
    print('Welcome aboard') 


print("Start")
# greet_user() am dros programa errors mogvcems
#  radgan nameshi araferia
greet_user("Saba")
greet_user("Nia")
print("Finish")
'''
'''
def greet_user(first_name, last_name): # am dros ukve programas veunebit
    # rom pirveli iqneba firstname da meore lastname rodesac
    # gamovidzaxebt mag. greet_user("Saba", "Esebua")     
    print(f'Hi {first_name} {last_name}!')
    print('Welcome aboard') 


print("Start")
greet_user("Saba", "Esebua")
print("Finish")
'''
# 29 Keyword Arguments
'''
def greet_user(first_name, last_name):    
    print(f'Hi {first_name} {last_name}!')
    print('Welcome aboard') 


print("Start")
greet_user("Saba", last_name="Esebua") # aseve shegvidzlia
# mivutitot pirdapir last_name da first_name rasnishnavs
#  magram konkretul am magalitze zevit mocemuli kodi jobs
# mokle da martivia..
print("Finish")
'''
# 30 Return Statement
'''
def square(number):
    return number * number

result = square(3)
print(result)


print(square(3)) # mxolod am kodit shegvidzlia davprintot
#rom ricxvi aris sami.. mokle da martivia
'''
# 31 Creating a Reusable Function / meoradi funqciis sheqmna
'''
def emoji_converter(message):
    words = message.split(" ")
    emojis = {
    ":)": "\N{winking face}",
    }
    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
    return output

message = input(">")
print(emoji_converter(message))   
'''
# 32 Exceptions / Gamonaklisi
'''
try:
    age = int(input('Age: ')) # int drosm xolod cifrebis
    # gamoeyneba sheileba 
    income = 20000
    risk = income / age
    print(age) 
except ZeroDivisionError: # anu 0 is errror..
    print('Age cannot be 0.') # tu zerodivision error miviget
    # daprintos rom 0 arsheileba da shemdegshi 0  armiutitos
except ValueError: # tu cifrebis magivrad miutita sityva
    print('Invalid value') # miutitos invalid value
'''
# 33  Comments
'''
# print sky is blue  . msgavsi komentari aris cudi vinaidan
# qvevit mocemulit mivxdebit da meore is rom
# tu shevcvlit sky is magalitad oceaneti 
# unda davbrundet da shevcvalot isic..
print("Sky is blue")  

# msgavs shemtxvevashi komentari aris kargi
# rom squareshi chawerili ricxvi 
# gamvravldeba tavis tavze.. 
def square(number):
return number * number    
'''
# 34 Classes
'''
class Point:  #klasis dros viyenebt did asos pirvelad..
    def move(self):
        print("move")

    def draw(self):
        print("draw")



point1 = Point() # gamodzaxebit Point() aseve martiv cvlads
# viyenebt ro gamodzaxebits dros point1 davwerot..         
point1.x = 10 # aseve shegvidzlia axlidan davamatot
# pint1.ze mnisvhnelobebi rom daprintvis dros
# amoagdos esec
point1.y = 20
print(point1.x)
point1.draw()

# aseve axal cvlads tu gavaigivebt Point tan
# es imas arnishnavs rom point1s msgavsad
# poin2.x printvis dros dagviwers 10's
# es point aris tavdapirvelad chawerili funqciashi
# draw da move anu..
point2 = Point() 
point2.x = 15
print(point2.x) 
'''
# 35 Constructors
'''
class Point:
    def __init__(self, x, y):
        self.x = x
        self.y = y
             
    def move(self):
        print("move")
              

    def draw(self):
        print("draw")

       
point = Point(10, 20)
point.x = 11
print(point.x)
'''
'''
class Person:
    def __init__(self, name):
        self.name = name
        

    def talk(self):
        print(f"Hi, I am {self.name}")


saba = Person("Saba Esebua")
saba.talk()

mate = Person("Lado Tkabladze")
mate.talk()
'''
# 36 Inheritance    
'''
class Mammal:    # dzudzumwovari
    def walk(self):
        print("walk")



class Dog(Mammal):
    def bark(self):
        print("bark")
    


class Cat(Mammal):
    def be_annoying(self):
        print("annoying")

cat1 = Cat()
cat1.be_annoying()

dog1 = Dog()
dog1.walk()
'''
# 37 Modules
'''      # importit im failshi myofi failis gadmogdeba shegidzlia
da archirdeba magalitad .py mititeba import converters.py nacvlad
import converters iqneba aseve is tvitonve gkarnaxobs
cvladebs da a.sh 
import converters
from converters import kg_to_lbs

kg_to_lbs(100)

print(converters.kg_to_lbs(68))
'''
'''
# utils py shi weria eseni.
# def find_max(numbers):          
#    max = numbers[0]
#    for number in numbers:
#        if number > max:
#            max = number
#    return max

# import utils
# utils.find_max()
from utils import find_max

numbers = [10, 3, 6, 2]
# max = find_max(numbers)
print(max(numbers))
'''
# 38 Packages
'''
# raime packages tu shevqnit magalitad saxelad 
# "ecommerce" da shignit iqneba python fili saxelad "shipping"
# chven shegvidzlia gamovidzaxot es file shemdeg nairad
# shemdeg ki rogorc zevit gamodzaxebaze iyo yvelaferi iseq iqenba
# importit..
#1
import ecommerce.shipping
ecommerce.shipping.calc_shipping()
# calc_shipping ki python shi iyo cvladis saxeli albat mixvdit
'''
'''
#2
from ecommerce.shipping import calc_shipping, calc_tax

calc_shipping()
'''
# 39 Generating Random Values
'''
import random


for i in range(3):
    print(random.random())
    print(random.randint(10, 20))
'''
'''
# randomze am otxi saxelidan gviprintavs liders amjamad
import random


members = ['Saba', 'Mari', 'Lado', 'Vero']
leader = random.choice(members)
print(leader)
'''
'''
# randomze roll shveba da amovardeba cifrebi
import random


class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        return first, second


dice = Dice()
print(dice.roll())        
'''
# 40 Files and Directories

# Absolute path
# c:\Program Files\Microsoft
# diski:file local:directory
# Releative path

'''
# vgebulobt aris tu ara am failshi magalitad "ecommerce"
# tu aris daprintavs pasuxs kis an aras
from pathlib import Path

path = Path("ecommerce") 
print(path.exists())
'''
'''
from pathlib import Path

path = Path()
#print(path.glob('*.*')) # amit vgebulobt yvela fails
# *.* yvelas 
# *.py mxolod python file da a.sh
for file in path.glob('*.py'): # daprintavs yvela py fails
    print(file) # am failshi
# xolo tu mxolod '*' chavwer daprintavs yvela failshi myof
# chanawers tu direqtors da yvelafers ra DD
'''    
# 41 Pypi and Pip 
# pypi.org 
# shegvidzlia davaintsalirot da gavigot rogor gamoviyenot
# uamravi tonobit masala
'''
'''

#42 Project 1: Automation with python
# Excel spreadsheets
'''
import openpyxl as xl 
from openpyxl.chart import BarChart, Reference


wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)
#print(cell.value)
#print(sheet.max_row)
'''
#for row in range(1, 4) # 1 2 3 not 4
# amit chven davpritanvt am failshi ramdeni xazia maqsimalurad
# sheuzgudavia anu programa ramdeni lainicaa imas dawers
# for row in range(1, sheet.max_row + 1):
#     print(row)
# amit amobechdavs meore zolidan yvela tanxas im rigshi 
# row, 3 nishnavs anu mesamedan anu mesame rigshi iyo 
# tanxa 2c, 3c, 4c eseni shesabamisad amoprintavs 
# shig chaweril tanxas  
#for row in range(2, sheet.max_row + 1):
#    cell = sheet.cell(row, 3)
#    print(cell.value)    
'''
# chven axla damatebit mivutotm cell mesame rigidan anu pricedan
# gaemravlebia 0.9 ze da migebuli ricxvebi daewera satitaod axal axal
# rowshi meotxe xazshi..
# shemdeg ki axali xlsx file vqmnit radgan 
# tu programa errors gamoiwvevs originali file ar daziandeba
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# import openxyl qvevit vwert amas
# from openpyxl.chart import BarChart, Reference
values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')
'''
# correct write project
# looks like a proffesional
'''
import openpyxl as xl 
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']


    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price


    values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)


    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
'''    
# 43 Project 2: Machine Learning with Python 
